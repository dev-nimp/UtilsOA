import openpyxl

wb = openpyxl.Workbook()
wb = openpyxl.load_workbook(filename='sample.xlsx')
sheets = wb.sheetnames

ws = wb[sheets[0]]
print(ws['A2'].value)
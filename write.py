from openpyxl import load_workbook
from configparser import ConfigParser
import pyodbc

# Конфигурация
parser = ConfigParser()
parser.read_file(open("conf.ini"))
FileLoad = parser['Config']['FileLoad']
ParametrsCell_1 = parser['Parametrs']['cell_1']
ParametrsCell_2 = parser['Parametrs']['cell_2']

# Соединение
cnxn = pyodbc.connect(
    r'DRIVER={SQL Server};'
    r'SERVER=DELOWEB-SERVER\SQLTEST;'
    r'DATABASE=DBTEST;'
    r'UID=sa;'
    r'PWD=1Nz3Z88Lc'
)

cursor = cnxn.cursor()
cursor.execute("SELECT * FROM dbo.AbUser")
row = cursor.fetchone()
# Проверка
print(row[0])
print(row[1])

# Работа с Excel
wb = load_workbook(FileLoad)
ws = wb.active
ws[ParametrsCell_1] = str(row[0])
ws[ParametrsCell_2] = int(row[1])

wb.save(FileLoad)